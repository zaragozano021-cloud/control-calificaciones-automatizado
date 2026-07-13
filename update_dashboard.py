import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, Reference

def generar_dashboard_academico(input_filename, output_filename):
    # 1. Cargar el libro de trabajo original
    wb = openpyxl.load_workbook(input_filename)
    ws_dash = wb["Dashboard"]
    ws_asist = wb["Control_Asistencia"]

    # 2. Limpieza completa de celdas y rangos mezclados antiguos en el Dashboard
    merged_ranges = list(ws_dash.merged_cells.ranges)
    for rng in merged_ranges:
        try:
            ws_dash.unmerge_cells(str(rng))
        except ValueError:
            pass

    for row in range(1, 45):
        for col in range(1, 35):
            cell = ws_dash.cell(row=row, column=col)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.font = Font(name="Segoe UI", size=11)

    # Forzar visibilidad de líneas de cuadrícula
    ws_dash.views.sheetView[0].showGridLines = True

    # 3. Estilos visuales compartidos (Paleta Ejecutiva / Slate)
    color_text_dark = "1E293B"
    color_text_muted = "64748B"
    color_accent = "0F766E"
    color_card_bg = "FFFFFF"
    color_card_border = "E2E8F0"

    font_title = Font(name="Segoe UI", size=15, bold=True, color=color_text_dark)
    font_subtitle = Font(name="Segoe UI", size=9, italic=True, color=color_text_muted)
    font_card_lbl = Font(name="Segoe UI", size=9, bold=True, color=color_text_muted)
    font_card_val = Font(name="Segoe UI", size=22, bold=True, color=color_text_dark)

    fill_header = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    fill_accent = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    border_subtle = Border(
        left=Side(style='thin', color=color_card_border),
        right=Side(style='thin', color=color_card_border),
        top=Side(style='thin', color=color_card_border),
        bottom=Side(style='thin', color=color_card_border)
    )

    # 4. Estructuración de Títulos en el Dashboard
    ws_dash["A2"] = "Dashboard de Rendimiento Académico"
    ws_dash["A2"].font = font_title
    ws_dash["A3"] = "Control de Calificaciones y Asistencia"
    ws_dash["A3"].font = font_subtitle

    # 5. Helper para construir tarjetas KPI Globales superiores
    def make_card(ws, start_c, end_c, label, formula, fmt=None):
        ws.merge_cells(start_row=4, start_column=start_c, end_row=4, end_column=end_c)
        ws.merge_cells(start_row=5, start_column=start_c, end_row=6, end_column=end_c)
        
        l_cell = ws.cell(row=4, column=start_c, value=label)
        l_cell.font = font_card_lbl
        l_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        v_cell = ws.cell(row=5, column=start_c, value=formula)
        v_cell.font = font_card_val
        v_cell.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            v_cell.number_format = fmt
            
        for r in range(4, 7):
            for c in range(start_c, end_c + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill(start_color=color_card_bg, end_color=color_card_bg, fill_type="solid")
                cell.border = border_subtle

    make_card(ws_dash, 2, 3, "TOTAL ESTUDIANTES", "=COUNTA(Base_Notas!B2:B100)")
    make_card(ws_dash, 4, 5, "PROMEDIO GENERAL CURSO", "=AVERAGE(Base_Notas!U2:U100)", "0.00")
    make_card(ws_dash, 6, 7, "ESTUDIANTES APROBADOS", '=COUNTIF(Base_Notas!V2:V100, "Cumple Objetivos")')
    make_card(ws_dash, 8, 8, "REPROBADOS", '=COUNTIF(Base_Notas!V2:V100, "No Cumple")')

    # 6. Módulo: Filtro Selector Individual por Alumno
    ws_dash.merge_cells("A9:E9")
    lbl_f = ws_dash["A9"]
    lbl_f.value = "FILTRO INDIVIDUAL POR ESTUDIANTE"
    lbl_f.font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    lbl_f.fill = fill_header
    lbl_f.alignment = Alignment(horizontal="center", vertical="center")

    ws_dash["A11"] = "Seleccione Alumno:"
    ws_dash["A11"].font = Font(name="Segoe UI", size=10, bold=True, color=color_accent)

    ws_dash.merge_cells("A12:D12")
    selector = ws_dash["A12"]
    selector.value = "BARCENAS PORTILLA ANDRES FELIPE"  # Inicializador por defecto
    selector.font = Font(name="Segoe UI", size=11, bold=True, color="000000")
    selector.fill = fill_accent
    selector.alignment = Alignment(horizontal="left", vertical="center")

    for col in range(1, 5):
        ws_dash.cell(row=12, column=col).border = border_subtle

    # Regla de Validación de Datos para el buscador dinámico
    dv_selector = DataValidation(type="list", formula1="Base_Notas!$B$2:$B$100", allow_blank=True)
    dv_selector.error = 'El estudiante no está en la lista de calificaciones'
    dv_selector.errorTitle = 'Estudiante no válido'
    ws_dash.add_data_validation(dv_selector)
    dv_selector.add(selector)

    # 7. Tabla de Cortes y Fórmulas VLOOKUP del Alumno Seleccionado
    hdrs = ["Periodo / Corte", "Porcentaje", "Nota Obtenida", "Estado del Corte"]
    for idx, h in enumerate(hdrs, start=2):
        cell = ws_dash.cell(row=14, column=idx, value=h)
        cell.font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_subtle

    # Índices exactos corregidos con base en la estructura real de Base_Notas
    cortes_mapeo = [
        ("Primer Corte", "30%", 6, 7),
        ("Segundo Corte", "30%", 12, 13),
        ("Tercer Corte", "40%", 18, 19),
        ("Definitiva Semestre", "100%", 20, 21)
    ]

    for i, (name, pct, idx_nota, idx_est) in enumerate(cortes_mapeo, start=15):
        ws_dash.cell(row=i, column=2, value=name).font = Font(name="Segoe UI", size=10, bold=True if i==18 else False)
        ws_dash.cell(row=i, column=3, value=pct).alignment = Alignment(horizontal="center")
        
        n_cell = ws_dash.cell(row=i, column=4, value=f"=VLOOKUP($A$12, Base_Notas!$B$2:$V$100, {idx_nota}, FALSE)")
        n_cell.font = Font(name="Segoe UI", size=10, bold=True)
        n_cell.number_format = '0.00'
        n_cell.alignment = Alignment(horizontal="center")
        
        e_cell = ws_dash.cell(row=i, column=5, value=f"=VLOOKUP($A$12, Base_Notas!$B$2:$V$100, {idx_est}, FALSE)")
        e_cell.font = Font(name="Segoe UI", size=10, bold=True)
        e_cell.alignment = Alignment(horizontal="center")
        
        for col in range(2, 6):
            ws_dash.cell(row=i, column=col).border = border_subtle

    # Fila de Inasistencias individuales abajo del bloque de notas
    ws_dash.merge_cells("B19:C19")
    ws_dash.cell(row=19, column=2, value="Total Faltas Alumno:").font = Font(name="Segoe UI", size=10, bold=True)
    f_cell = ws_dash.cell(row=19, column=4, value="=VLOOKUP($A$12, Control_Asistencia!$B$2:$AK$100, 34, FALSE)")
    f_cell.font = Font(name="Segoe UI", size=10, bold=True)
    f_cell.alignment = Alignment(horizontal="center")
    ae_cell = ws_dash.cell(row=19, column=5, value="=VLOOKUP($A$12, Control_Asistencia!$B$2:$AK$100, 36, FALSE)")
    ae_cell.font = Font(name="Segoe UI", size=10, bold=True)
    ae_cell.alignment = Alignment(horizontal="center")

    for col in range(2, 6):
        ws_dash.cell(row=19, column=col).border = border_subtle

    # 8. Módulo Fijo: Estudiantes con Alerta de Reprobación (< 3.0)
    ws_dash.merge_cells("B21:E21")
    lbl_rep = ws_dash["B21"]
    lbl_rep.value = "ESTUDIANTES REPROBADOS (NOTA < 3.0)"
    lbl_rep.font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    lbl_rep.fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
    lbl_rep.alignment = Alignment(horizontal="center", vertical="center")

    reprobados_data = [
        ("Gómez Jaramillo Diana Carolina", 2.36),
        ("Zapata Ossa Gabriela", 2.72)
    ]
    for idx, (name, score) in enumerate(reprobados_data, start=22):
        ws_dash.cell(row=idx, column=2, value=name).font = Font(name="Segoe UI", size=10, bold=True, color=color_text_dark)
        ws_dash.cell(row=idx, column=3, value="Nota Final:").font = Font(name="Segoe UI", size=9, color=color_text_muted)
        cs = ws_dash.cell(row=idx, column=4, value=score)
        cs.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
        cs.number_format = '0.00'
        ws_dash.cell(row=idx, column=5, value="No Cumple").font = Font(name="Segoe UI", size=9, italic=True, color="991B1B")
        for col in range(2, 6):
            ws_dash.cell(row=idx, column=col).border = border_subtle

    # 9. Tabla Específica Oculta / de Apoyo para RAPs (Columnas AE, AF, AG)
    ws_dash["AE1"] = "Componente"
    ws_dash["AF1"] = "Promedio Grupo"
    ws_dash["AG1"] = "Estudiante Seleccionado"
    
    raps_labels = ["RAP 1: Fundamentos (C1)", "RAP 2: Aplicación (C2)", "RAP 3: Integración (C3)"]
    formulas_grupo = ["=AVERAGE(Base_Notas!G2:G100)", "=AVERAGE(Base_Notas!M2:M100)", "=AVERAGE(Base_Notas!S2:S100)"]
    formulas_alumno = [
        "=VLOOKUP($A$12, Base_Notas!$B$2:$V$100, 6, FALSE)",
        "=VLOOKUP($A$12, Base_Notas!$B$2:$V$100, 12, FALSE)",
        "=VLOOKUP($A$12, Base_Notas!$B$2:$V$100, 18, FALSE)"
    ]
    for idx, (lbl, f_g, f_a) in enumerate(zip(raps_labels, formulas_grupo, formulas_alumno), start=2):
        ws_dash.cell(row=idx, column=31, value=lbl)
        ws_dash.cell(row=idx, column=32, value=f_g).number_format = '0.00'
        ws_dash.cell(row=idx, column=33, value=f_a).number_format = '0.00'

    # 10. Construcción Limpia de Gráficos Nativos
    ws_dash._charts.clear()

    # Gráfico A: Rendimiento Individual por Corte (Barras)
    chart_ind = BarChart()
    chart_ind.type = "col"
    chart_ind.style = 10
    chart_ind.title = "RENDIMIENTO INDIVIDUAL POR CORTE ACADÉMICO"
    chart_ind.y_axis.title = "Calificación (0.0 - 5.0)"
    chart_ind.width = 11
    chart_ind.height = 6.2
    chart_ind.legend = None

    d_ref = Reference(ws_dash, min_col=4, min_row=14, max_row=17) 
    c_ref = Reference(ws_dash, min_col=2, min_row=15, max_row=17) 
    chart_ind.add_data(d_ref, titles_from_data=True)
    chart_ind.set_categories(c_ref)
    ws_dash.add_chart(chart_ind, "K4")

    # Gráfico B: Análisis Avanzado Comparativo RAPs (Líneas Multiserie)
    chart_compare = LineChart()
    chart_compare.title = "ANÁLISIS COMPARATIVO: ESTUDIANTE VS PROMEDIO DEL GRUPO POR RAP"
    chart_compare.style = 13
    chart_compare.y_axis.title = "Calificación Académica"
    chart_compare.x_axis.title = "Resultados de Aprendizaje (RAP)"
    chart_compare.width = 19.5
    chart_compare.height = 7.5

    data_compare = Reference(ws_dash, min_col=32, max_col=33, min_row=1, max_row=4)
    cats_compare = Reference(ws_dash, min_col=31, min_row=2, max_row=4)
    chart_compare.add_data(data_compare, titles_from_data=True)
    chart_compare.set_categories(cats_compare)
    ws_dash.add_chart(chart_compare, "B25")

    # 11. REDISEÑO COMPLETO Y DIDÁCTICO DE ASISTENCIA
    # Sincronización mediante fórmulas dinámicas con Base_Notas
    for r in range(2, 100):
        ws_asist.cell(row=r, column=1, value=f"=IF(Base_Notas!A{r}<>\"\", Base_Notas!A{r}, \"\")")
        ws_asist.cell(row=r, column=2, value=f"=IF(Base_Notas!B{r}<>\"\", Base_Notas!B{r}, \"\")")

    # Estructura didáctica de celdas y validación de datos (0, 1, J)
    dv_asist = DataValidation(type="list", formula1='"0,1,J"', allow_blank=True)
    ws_asist.add_data_validation(dv_asist)

    # Paleta de Colores Pastel Didácticos
    fill_presente = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    fill_falta = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
    fill_justificado = PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid")

    font_presente = Font(name="Segoe UI", size=10, color="137333", bold=True)
    font_falta = Font(name="Segoe UI", size=10, color="C5221F", bold=True)
    font_justificado = Font(name="Segoe UI", size=10, color="B06000", bold=True)
    
    border_thin = Side(style='thin', color='CBD5E1')
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # Formateo de dimensiones y renderizado didáctico
    ws_asist.column_dimensions['A'].width = 15
    ws_asist.column_dimensions['B'].width = 38
    for col_idx in range(3, 35):
        ws_asist.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 6

    for row in range(2, 45):
        for col in range(3, 35):
            cell = ws_asist.cell(row=row, column=col)
            if cell.value is None:
                cell.value = 0
            val = str(cell.value)
            
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border
            
            if val == "1":
                cell.fill = fill_falta
                cell.font = font_falta
            elif val.upper() == "J":
                cell.fill = fill_justificado
                cell.font = font_justificado
            else:
                cell.fill = fill_presente
                cell.font = font_presente

    # Actualización de fórmulas lógicas de estado de asistencia
    for r in range(2, 100):
        ws_asist.cell(row=r, column=35, value=f"=IF(B{r}<>\"\", SUMIF(C{r}:AH{r}, 1), \"\")")
        ws_asist.cell(row=r, column=36, value=f"=IF(B{r}<>\"\", AI{r}/32, \"\")")
        ws_asist.cell(row=r, column=37, value=f"=IF(B{r}=\"\", \"\", IF(AJ{r}>0.2, \"Pérdida por Faltas\", \"Regular\"))")

    # Guardar el archivo definitivo
    wb.save(output_filename)
    print(f"¡Éxito! Archivo generado correctamente en: {output_filename}")

if __name__ == "__main__":
    # Define tus rutas aquí al ejecutar en tu entorno local o de CI/CD
    if __name__ == "__main__":
    # Asegúrate de dar 4 espacios (o una tecla TAB) al inicio de la siguiente línea:
    generar_dashboard_academico("Control_Calificaciones_Asistencia_Didactica.xlsx", "Control_Calificaciones_GitHub.xlsx")
